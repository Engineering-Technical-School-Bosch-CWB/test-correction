import math
import numpy as np
import openpyxl 
import cv2

def readExcelGabarito():
    gabarito = []
    pontuacao = []
    caminho = "ProcessoSeletivo.xlsx"

    df = openpyxl.load_workbook(caminho)
    
    sheet = df["Gabarito"]
    column = sheet.max_row

    for i in range(2, column + 1):
        valor = sheet.cell(row = i, column = 2)
        resposta = sheet.cell(row = i, column = 3)
        gabarito.append((valor.value, str(resposta.value)))
        pontuacao.append(sheet.cell(row = i, column = 1).value)

    return gabarito, pontuacao

def readExcelNome():
    names = []
    caminho = "ProcessoSeletivo.xlsx"

    df = openpyxl.load_workbook(caminho)
    
    sheet = df["Nomes"]
    coluna = sheet.max_row

    for i in range(2, coluna + 1):
        nome = sheet.cell(row = i, column = 1)
        names.append(nome.value)

    return names

def registerCandidate(candidate, race, answers, grade, answersWithOptions):
    caminho = "ProcessoSeletivo.xlsx"

    df = openpyxl.load_workbook(caminho)
    sheet = df["Registros"]
    exists, line = findCandidate(candidate)
    row = sheet.max_row + 1

    if (exists):
        row = line

    sheet.cell(row=row, column=1).value = candidate
    sheet.cell(row=row, column=2).value = grade
    sheet.cell(row=row, column=3).value = race

    _option = ['a','b','c','d','e']
    _answers = []
    for i in range(20):
        added = False
        for j in range(5):
            if answersWithOptions[i][j] != 0:
                _answers.append(_option[j])
                added = True
                break
        if(not added):
            _answers.append("-")
            
    for i in range(len(_answers)):
        sheet.cell(row=row, column=i + 4).value = _answers[i]

    df.save("ProcessoSeletivo.xlsx")
    
    return "Success", 200

def findCandidate(candidate):
    caminho = "ProcessoSeletivo.xlsx"
    exists = False
    line = 1

    df = openpyxl.load_workbook(caminho)
    sheet = df["Registros"]
    linha = sheet.max_row 

    for i in range(1, linha + 1):
        line = i
        nome = sheet.cell(row = i, column = 1)
        if (nome.value == candidate):
            exists = True
            break
    
    return exists, line

def warp(dict, img, imgWidth, imgHeight, inverse):
    if inverse:
        pts2 = np.float32([[dict[326][0], dict[326][1]], [dict[683][0], dict[683][1]],[dict[779][0], dict[779][1]], [dict[856][0], dict[856][1]]])
        pts1 = np.float32([[0, 0],[imgWidth, 0],[imgWidth, imgHeight], [0, imgHeight]])
    else:
        pts1 = np.float32([[dict[326][0], dict[326][1]], [dict[683][0], dict[683][1]],[dict[779][0], dict[779][1]], [dict[856][0], dict[856][1]]])
        pts2 = np.float32([[0, 0],[imgWidth, 0],[imgWidth, imgHeight], [0, imgHeight]])
    M = cv2.getPerspectiveTransform(pts1,pts2)
    dst = cv2.warpPerspective(img, M, (imgWidth, imgHeight))
    return dst

# 

def splitBoxes(img, questions, options, questionsOverlapMargin):
    """
        A method to verify if user marked an option by question\n
        options = option of question (A,B,C,D,E...)\n
        question = question of test (1 ~ 20)

    """
    total = 0
    
    rows = np.vsplit(img, questions)
    boxes = []
    for r in rows:
        cols = np.hsplit(r, options)
        for box in cols:
            total += np.count_nonzero(box[questionsOverlapMargin:box.shape[0]-questionsOverlapMargin, questionsOverlapMargin:box.shape[1]-questionsOverlapMargin])
            boxes.append(box[questionsOverlapMargin:box.shape[0]-questionsOverlapMargin, questionsOverlapMargin:box.shape[1]-questionsOverlapMargin])

    return boxes, int(total/(questions*options))

def giveGrades(answers, rigthAnswers):
    grade = 0
    
    answersList = []
    for a in range(len(answers)):
        if sum(answers[a]) != 1 or (ord(rigthAnswers[a][1]) - 65) != np.argmax(answers[a]):
            answersList.append(0)
        else:
            answersList.append(1)
            grade += rigthAnswers[a][0]
            answers[a][np.argmax(answers[a])] = -1

    return answers, grade, answersList

def showAnswers(img, answers, questions, choices):
    optionH = int(img.shape[1]/questions)
    optionW = int(img.shape[0]/choices)

    for currentRow in range(questions):
        if sum(answers[currentRow]) == -1:
            answer = np.argmin(answers[currentRow])
            cX = int((answer*optionW) + optionW / 2)
            cY = int((currentRow * optionH) + optionH / 2)
            cv2.ellipse(img, (cX,cY), (50, 15), 0, 0, 360, (0, 255, 0), -1)
        elif (sum(answers[currentRow]) == 1):
            answer = np.argmax(answers[currentRow])
            cX = int((answer*optionW) + optionW / 2)
            cY = int((currentRow * optionH) + optionH / 2)
            cv2.ellipse(img, (cX,cY), (50, 15), 0, 0, 360, (0, 0, 255), -1)

    return img

#! ================== making debug image to view data ============================

def showDebug(images, width = 1080, height = 720, title= "debug"):
    background = np.zeros((height, width, 3), dtype=np.uint8)
    img_height = int(height / 2)
    img_width = int(width / 4)
    for i in range(len(images)):
        if len(images[i].shape) == 2:
            images[i] = cv2.cvtColor(images[i], cv2.COLOR_GRAY2BGR)
        images[i] = cv2.resize(images[i], (img_width, img_height))
        line = math.floor(i / 4)
        column = i % 4
        background[
            line * img_height : (line + 1) * img_height,
            column * img_width: img_width * (column + 1)
            ] = images[i]
        cv2.imshow(title, background)


def listAvailableCams(max_cameras = 10):
    cameras_disponiveis = []
    for index in range(max_cameras):
        cap = cv2.VideoCapture(index, cv2.CAP_DSHOW)  # No Windows, use CAP_DSHOW
        if cap is not None and cap.isOpened():
            cameras_disponiveis.append(index)
            cap.release()
    return cameras_disponiveis